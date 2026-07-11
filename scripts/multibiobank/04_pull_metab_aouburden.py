#!/usr/bin/env python3
"""Parse UKB metabolomics credible sets (Nat Genet 2025 ST5) and AoU 392k burden G2T hits.
Metab: s41588-025-02355-3 MOESM4 ST5 (43,322 CS-NMR pairs, SuSiE, hg38, 249 traits).
AoU burden: medRxiv 2026.05.08.26350964 supplement media-2.xlsx SD3/SD5/SD7 (G2T resource, sig only).
"""
import openpyxl, polars as pl

# ---- UKB metabolomics ST5 ----
wb=openpyxl.load_workbook("metab_st.xlsx", read_only=True); ws=wb["ST5"]
rows=list(ws.iter_rows(values_only=True)); hdr_idx=2
header=[str(c).strip() if c is not None else f"c{j}" for j,c in enumerate(rows[hdr_idx])]
data=[["" if c is None else str(c) for c in r] for r in rows[hdr_idx+1:] if any(c is not None for c in r)]
df=pl.DataFrame(data, schema=header, orient="row")
u=df.with_columns([
    (pl.col("chrom").str.replace(r"^chr","")+"_"+pl.col("pos.hg38")+"_"+pl.col("allele0")+"_"+pl.col("allele1")).alias("variant_hg38"),
    pl.col("chrom").str.replace(r"^chr","").alias("chromosome"),
    pl.col("pos.hg38").cast(pl.Int64,strict=False).alias("position"),
    pl.col("allele0").alias("ref"), pl.col("allele1").alias("alt"), pl.col("pheno").alias("trait"),
    pl.lit("UKB_metabolomics").alias("study"), pl.col("pip").cast(pl.Float64,strict=False).alias("PIP"),
    (pl.col("pheno")+"|"+pl.col("region")).alias("cs_id"), pl.lit(True).alias("cs_membership"),
    pl.col("R2").cast(pl.Float64,strict=False).alias("r2_to_lead"),
    pl.lit("multi_ancestry_UKB").alias("ancestry"), pl.lit("imputation").alias("assay_basis"),
    pl.lit("hg38").alias("build"), pl.lit("UKB_metabolomics").alias("source"),
    pl.col("beta_joint").cast(pl.Float64,strict=False).alias("beta"), pl.col("SYMBOL.lead").alias("gene")])
SCHEMA=["variant_hg38","chromosome","position","ref","alt","trait","study","PIP","cs_id","cs_membership","r2_to_lead","ancestry","assay_basis","build","source","beta","gene"]
u.select(SCHEMA).filter(pl.col("PIP").is_not_null()).write_parquet("ukb_metabolomics_credible_sets.parquet")

# ---- AoU 392k burden G2T ----
wb2=openpyxl.load_workbook("aou_supp.xlsx", read_only=True)
def read_sheet(sn, hr=2):
    ws=wb2[sn]; rr=list(ws.iter_rows(values_only=True))
    h=[str(c).strip() if c is not None else f"c{j}" for j,c in enumerate(rr[hr-1])]
    d=[["" if c is None else str(c) for c in r] for r in rr[hr:] if any(c is not None for c in r)]
    return pl.DataFrame(d, schema=h, orient="row")
sd3,sd5,sd7=read_sheet("Supplementary Data 3"),read_sheet("Supplementary Data 5"),read_sheet("Supplementary Data 7")
sd3u=sd3.select([pl.col("gene_symbol").alias("gene"),pl.col("gene_id"),pl.col("phenoname").alias("phenotype_id"),pl.col("description").alias("trait"),pl.col("category"),pl.col("annotation"),pl.col("meta_BETA_Burden").cast(pl.Float64,strict=False).alias("beta_burden"),pl.col("meta_p").cast(pl.Float64,strict=False).alias("pval_burden"),pl.col("similarity_group").alias("ancestry_group"),pl.lit("SD3_216pLoF").alias("supp_table")]).unique(subset=["gene","phenotype_id","ancestry_group"])
sd5u=sd5.select([pl.col("gene_symbol").alias("gene"),pl.col("gene_id"),pl.col("phenoname").alias("phenotype_id"),pl.col("description").alias("trait"),pl.col("category"),pl.col("annotation"),pl.col("BETA_Burden").cast(pl.Float64,strict=False).alias("beta_burden"),pl.col("Pvalue_Burden").cast(pl.Float64,strict=False).alias("pval_burden"),pl.lit("META").alias("ancestry_group"),pl.lit("SD5_44crossdomain").alias("supp_table")])
sd7u=sd7.select([pl.col("gene_symbol").alias("gene"),pl.col("gene_id"),pl.col("phenoname").alias("phenotype_id"),pl.lit(None,dtype=pl.Utf8).alias("trait"),pl.col("category"),pl.lit("pLoF").alias("annotation"),pl.lit(None,dtype=pl.Float64).alias("beta_burden"),pl.col("META_Pvalue_Burden").cast(pl.Float64,strict=False).alias("pval_burden"),pl.lit("META").alias("ancestry_group"),pl.lit("SD7_193novel").alias("supp_table")])
pl.concat([sd3u,sd5u,sd7u],how="vertical_relaxed").with_columns([pl.lit("AoU_392k_AllbyAll").alias("source"),pl.lit("hg38").alias("build"),pl.lit("burden_pLoF").alias("evidence_type")]).write_parquet("aou_burden_g2t_sig.parquet")
